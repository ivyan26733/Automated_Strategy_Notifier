#!/usr/bin/env python3
"""Create an annual-return heatmap from NIFTY sectoral index daily histories.

Input files are the daily Yahoo histories downloaded to ``sector_data`` by
``fetch_yahoo_stock_history.py``.  The sector definitions are NIFTY sectoral
benchmarks; annual return = final close in a calendar year / final close in
the preceding calendar year - 1.
"""

from __future__ import annotations

import argparse
import csv
import os
from collections import defaultdict
from pathlib import Path

# Keep Matplotlib's generated font cache inside the project, where the script
# has write access even in restricted environments.
os.environ.setdefault("MPLCONFIGDIR", str(Path(__file__).resolve().parent / ".matplotlib"))

import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


SECTORS = {
    "Auto": "_CNXAUTO_daily_history.csv",
    "Bank": "_NSEBANK_daily_history.csv",
    "Financial Services": "NIFTY_FIN_SERVICE.NS_daily_history.csv",
    "FMCG": "_CNXFMCG_daily_history.csv",
    "IT": "_CNXIT_daily_history.csv",
    "Media": "_CNXMEDIA_daily_history.csv",
    "Metal": "_CNXMETAL_daily_history.csv",
    "Pharma": "_CNXPHARMA_daily_history.csv",
    "Realty": "_CNXREALTY_daily_history.csv",
    "Energy": "_CNXENERGY_daily_history.csv",
}

SECTOR_COLORS = {
    "Auto": "F39C12",
    "Bank": "1F4E78",
    "Financial Services": "4F81BD",
    "FMCG": "2E8B57",
    "IT": "00A6D6",
    "Media": "7F8C8D",
    "Metal": "E74C3C",
    "Pharma": "8E44AD",
    "Realty": "E91E63",
    "Energy": "6B8E23",
}


def annual_returns(history_dir: Path) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for sector, filename in SECTORS.items():
        path = history_dir / filename
        if not path.exists():
            print(f"Skipping {sector}: missing {path}")
            continue
        closes: dict[int, tuple[str, float]] = {}
        with path.open(newline="", encoding="utf-8") as file:
            for row in csv.DictReader(file):
                try:
                    date, close = row["Date"], float(row["Close"])
                    year = int(date[:4])
                except (KeyError, TypeError, ValueError):
                    continue
                # Files are chronological, so retain the final trading close.
                closes[year] = (date, close)
        for year in sorted(closes):
            if year - 1 not in closes:
                continue
            previous_date, previous_close = closes[year - 1]
            date, close = closes[year]
            # A yearly heatmap should contain only completed calendar years.
            if date[5:7] != "12":
                continue
            records.append(
                {
                    "Year": year,
                    "Sector": sector,
                    "Return (%)": (close / previous_close - 1) * 100,
                    "Start date": previous_date,
                    "Start close": previous_close,
                    "End date": date,
                    "End close": close,
                }
            )
    return records


def write_csv(records: list[dict[str, object]], path: Path) -> None:
    fields = ["Year", "Rank", "Sector", "Return (%)", "Start date", "Start close", "End date", "End close"]
    ranked = rank_records(records)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(ranked)


def write_excel(records: list[dict[str, object]], path: Path) -> None:
    """Write a colour-formatted workbook; CSV cannot preserve this formatting."""
    ranked = rank_records(records)
    workbook = Workbook()
    heatmap = workbook.active
    heatmap.title = "Yearly Heatmap"
    heatmap.sheet_view.showGridLines = False
    years = sorted({int(record["Year"]) for record in ranked})
    max_rank = max(int(record["Rank"]) for record in ranked)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    heatmap.cell(1, 1, "Rank")
    for index, year in enumerate(years, start=2):
        heatmap.cell(1, index, year)
    for cell in heatmap[1]:
        cell.fill, cell.font, cell.alignment = header_fill, Font(color="FFFFFF", bold=True), Alignment(horizontal="center")
    grid = {(int(record["Year"]), int(record["Rank"])): record for record in ranked}
    for rank in range(1, max_rank + 1):
        heatmap.cell(rank + 1, 1, rank)
        for index, year in enumerate(years, start=2):
            record = grid.get((year, rank))
            cell = heatmap.cell(rank + 1, index)
            if record:
                sector = str(record["Sector"])
                cell.value = f"{sector}\n{float(record['Return (%)']):+.1f}%"
                cell.fill = PatternFill("solid", fgColor=SECTOR_COLORS[sector])
                cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    heatmap.column_dimensions["A"].width = 8
    for column in range(2, len(years) + 2):
        heatmap.column_dimensions[chr(64 + column)].width = 18
    for row in range(2, max_rank + 2):
        heatmap.row_dimensions[row].height = 38
    heatmap.freeze_panes = "B2"

    raw = workbook.create_sheet("Annual Returns")
    fields = ["Year", "Rank", "Sector", "Return (%)", "Start date", "Start close", "End date", "End close"]
    raw.append(fields)
    for record in ranked:
        raw.append([record[field] for field in fields])
    for cell in raw[1]:
        cell.fill, cell.font = header_fill, Font(color="FFFFFF", bold=True)
    raw.freeze_panes = "A2"
    raw.auto_filter.ref = raw.dimensions

    for sector in SECTORS:
        sheet = workbook.create_sheet(sector[:31])
        sheet.sheet_properties.tabColor = SECTOR_COLORS[sector]
        sheet.append(["Year", "Return (%)", "Start date", "Start close", "End date", "End close"])
        for record in ranked:
            if record["Sector"] == sector:
                sheet.append([record[field] for field in ["Year", "Return (%)", "Start date", "Start close", "End date", "End close"]])
        for cell in sheet[1]:
            cell.fill, cell.font = PatternFill("solid", fgColor=SECTOR_COLORS[sector]), Font(color="FFFFFF", bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in "ABCDEF":
            sheet.column_dimensions[column].width = 16
    workbook.save(path)


def rank_records(records: list[dict[str, object]]) -> list[dict[str, object]]:
    by_year: dict[int, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        by_year[int(record["Year"])].append(record.copy())
    ranked: list[dict[str, object]] = []
    for year, entries in by_year.items():
        for rank, entry in enumerate(sorted(entries, key=lambda item: float(item["Return (%)"]), reverse=True), start=1):
            entry["Rank"] = rank
            entry["Return (%)"] = round(float(entry["Return (%)"]), 2)
            ranked.append(entry)
    return sorted(ranked, key=lambda item: (int(item["Year"]), int(item["Rank"])))


def draw_heatmap(records: list[dict[str, object]], path: Path) -> None:
    ranked = rank_records(records)
    years = sorted({int(record["Year"]) for record in ranked})
    max_rank = max(int(record["Rank"]) for record in ranked)
    grid = {(int(record["Year"]), int(record["Rank"])): record for record in ranked}
    returns = [float(record["Return (%)"]) for record in ranked]
    limit = max(20, max(abs(value) for value in returns))
    cmap = plt.colormaps["RdYlGn"]
    norm = Normalize(vmin=-limit, vmax=limit)

    figure, axis = plt.subplots(figsize=(max(12, len(years) * 0.85), max(7, max_rank * 0.55 + 1.5)))
    for x, year in enumerate(years):
        for rank in range(1, max_rank + 1):
            record = grid.get((year, rank))
            color = "#f1f3f5" if record is None else cmap(norm(float(record["Return (%)"])))
            axis.add_patch(Rectangle((x, max_rank - rank), 1, 1, facecolor=color, edgecolor="white", linewidth=1.2))
            if record:
                return_value = float(record["Return (%)"])
                text_color = "white" if abs(return_value) > limit * 0.62 else "#1f2933"
                label = str(record["Sector"]).replace(" ", "\n")
                axis.text(x + 0.5, max_rank - rank + 0.63, label, ha="center", va="center", fontsize=7.5, color=text_color, weight="bold", linespacing=0.9)
                axis.text(x + 0.5, max_rank - rank + 0.32, f"{return_value:+.1f}%", ha="center", va="center", fontsize=8, color=text_color)

    axis.set_xlim(0, len(years))
    axis.set_ylim(0, max_rank)
    axis.set_xticks([index + 0.5 for index in range(len(years))], years, fontsize=9)
    axis.set_yticks([])
    axis.tick_params(top=True, labeltop=True, bottom=False, labelbottom=False)
    for spine in axis.spines.values():
        spine.set_visible(False)
    axis.set_title("NIFTY Sectoral Indices — Calendar-Year Return Heatmap", pad=28, weight="bold", fontsize=15)
    axis.text(0, -0.55, "Each column is ranked best-to-worst by annual price return. Return uses the final trading close of each calendar year.", fontsize=9, color="#4b5563")
    colorbar = figure.colorbar(plt.cm.ScalarMappable(norm=norm, cmap=cmap), ax=axis, pad=0.015, shrink=0.8)
    colorbar.set_label("Calendar-year return (%)")
    figure.tight_layout()
    figure.savefig(path, dpi=220, bbox_inches="tight")


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a yearly NIFTY sectoral-index heatmap.")
    parser.add_argument("--history-dir", type=Path, default=Path("sector_data"))
    parser.add_argument("--output-dir", type=Path, default=Path("heatmap_output"))
    args = parser.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    records = annual_returns(args.history_dir)
    if not records:
        raise SystemExit("No usable sector-index history was found.")
    write_excel(records, args.output_dir / "nse_sector_heatmap.xlsx")
    draw_heatmap(records, args.output_dir / "nse_sector_yearly_heatmap.png")
    csv_created = True
    try:
        write_csv(records, args.output_dir / "nse_sector_annual_returns.csv")
    except PermissionError:
        csv_created = False
        print("Could not update nse_sector_annual_returns.csv because it is open in another program.")
    print(f"Created {args.output_dir / 'nse_sector_heatmap.xlsx'}")
    print(f"Created {args.output_dir / 'nse_sector_yearly_heatmap.png'}")
    if csv_created:
        print(f"Created {args.output_dir / 'nse_sector_annual_returns.csv'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
